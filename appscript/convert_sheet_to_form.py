"""
convert_sheet_to_form.py
========================
Converts an Excel file (or Google Sheet, or exported PDF) with {{tags}}
into a fillable AcroForm PDF.

MODES
-----
A) From Excel (simplest — no auth needed):
     py convert_sheet_to_form.py template.xlsx  output_form.pdf
     py convert_sheet_to_form.py template.xlsx  output_form.pdf  --sheet-name cover

B) From Google Sheets (requires Google auth):
     py convert_sheet_to_form.py --sheet-id SHEET_ID output_form.pdf
     py convert_sheet_to_form.py --sheet-id SHEET_ID --sheet-name cover output_form.pdf

C) From a locally exported PDF:
     py convert_sheet_to_form.py input.pdf output_form.pdf

Then upload:
     py upload_template.py

TAG SYNTAX in Excel/Sheet cells:
  {{txt:fieldName}}   → TextField  (single line)
  {{ml:fieldName}}    → TextField  (multiline)
  {{chk:fieldName}}   → Checkbox
  {{sig:fieldName}}   → Signature field (empty box)
  {{date:fieldName}}  → Date TextField

REQUIREMENTS:
  py -m pip install pymupdf
  (Excel mode on Windows uses win32com — requires Microsoft Excel to be installed)
  (Google Sheets mode: py -m pip install google-auth google-api-python-client)
"""

import sys
import json
import re
import io
import argparse
import tempfile
import os
import fitz  # PyMuPDF

# ── Tag patterns ────────────────────────────────────────────────────────────
# Typed:  {{txt:fieldName}}           → single-line, left
#         {{txt:fieldName:c}}         → single-line, center
#         {{txt:fieldName:r}}         → single-line, right
#         {{ml:fieldName}}            → multiline, auto-height from border
#         {{ml:fieldName:3}}          → multiline, 3 lines tall
#         {{ml:fieldName:3:1.5}}      → multiline, 3 lines (spacing stored, not rendered)
#         {{ml:fieldName:3:1.5:c}}    → multiline, 3 lines, center (spacing stored only)
#         {{ml:fieldName:1.5}}        → multiline, auto-height (spacing stored only)
#         {{ml:fieldName:c}}          → multiline, auto-height, center
# Plain:  {{fieldName}}               → txt, left
#
# Segment order (all optional): :N (int lines) :S (float spacing) :align (l/c/r)
# Distinguishable because: integer=lines, decimal=spacing, letter=alignment
TAG_TYPED = re.compile(
    r'\{\{(txt|ml|chk|sig|date|photo):([^}:]+)'
    r'(?::(\d+))?'           # group 3: line count  (integer only)
    r'(?::(\d+\.\d+))?'     # group 4: line spacing (must have decimal point)
    r'(?::([lcrLCR]))?'      # group 5: alignment
    r'\}\}'
)
TAG_PLAIN = re.compile(r'\{\{([a-zA-Z_][a-zA-Z0-9_]*)(?::([lcrLCR]))?\}\}')

_ALIGN_MAP = {'l': 0, 'L': 0, 'c': 1, 'C': 1, 'r': 2, 'R': 2}

# PDF viewers (Acrobat, Chrome, Edge) render AcroForm multiline text at roughly
# 1.25× the font size per line.  We use this constant to convert a line count
# into a pixel-accurate field height.  AcroForm does NOT support custom line
# spacing — the :S segment is accepted in the tag syntax and stored in the
# field-map JSON for reference, but it has no effect on the rendered PDF.
_VIEWER_LINE_HEIGHT = 1.25   # matches Acrobat / Chrome PDF viewer rendering
_VIEWER_LINE_PADDING = 4     # extra vertical padding (pts) for top+bottom

def parse_tag(text: str):
    """
    Returns (tag_type, tag_name, lines, spacing, explicit_align) or None.
      lines:         int   — number of text lines for box sizing (None = auto from border)
      spacing:       float — stored in field-map only; PDF viewers ignore it
      explicit_align: 0=left, 1=center, 2=right, None=auto-detect
    """
    m = TAG_TYPED.search(text)
    if m:
        lines    = int(m.group(3))   if m.group(3) else None
        spacing  = float(m.group(4)) if m.group(4) else None
        explicit = _ALIGN_MAP.get(m.group(5)) if m.group(5) else None
        return m.group(1), m.group(2), lines, spacing, explicit
    m = TAG_PLAIN.search(text)
    if m:
        explicit = _ALIGN_MAP.get(m.group(2)) if m.group(2) else None
        return 'txt', m.group(1), None, None, explicit
    return None


# ── Excel geometry helpers ──────────────────────────────────────────────────
# Used to convert Excel column widths / row heights to PDF points so that
# merged-cell boundaries can be determined directly from the workbook geometry
# rather than being inferred from PDF drawing paths (which is unreliable).
#
# Excel column widths are stored in "character units" (MDW-based).  For the
# default font (Calibri 11pt at 96 dpi) the Max Digit Width ≈ 7 px.
# Conversion: px = int(w * MDW + 5);  pt = px * 72 / 96
_EXCEL_MDW_PX = 7          # pixels per character-width unit (Calibri 11pt @ 96dpi)
_EXCEL_DEFAULT_COL_W = 8.0 # character units (Excel default)
_EXCEL_DEFAULT_ROW_H = 12.75  # points (Excel default)


def _excel_col_width_pt(ws, col_idx: int) -> float:
    """Width of column col_idx (1-based) in points."""
    try:
        import openpyxl.utils
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cd = ws.column_dimensions.get(col_letter)
        w = float(cd.width) if (cd and cd.width) else _EXCEL_DEFAULT_COL_W
    except Exception:
        w = _EXCEL_DEFAULT_COL_W
    px = int(w * _EXCEL_MDW_PX + 5)
    return px * 72.0 / 96.0


def _excel_row_height_pt(ws, row_idx: int) -> float:
    """Height of row row_idx (1-based) in points."""
    try:
        rd = ws.row_dimensions.get(row_idx)
        if rd and rd.height:
            return float(rd.height)
    except Exception:
        pass
    return _EXCEL_DEFAULT_ROW_H


def build_excel_tag_rects(xlsx_path: str, sheet_name: str = None) -> dict:
    """
    Open the workbook with openpyxl, find every {{tag}} cell, look up its
    merged range, and return the cell's bounding box in cumulative Excel points.

    Returns {tag_text: (x0_pt, y0_pt, x1_pt, y1_pt)} measured from the
    sheet origin (column 1, row 1).  These coordinates are later calibrated
    to PDF page space using the tag-text span positions as anchors.
    """
    try:
        import openpyxl
        import openpyxl.utils
    except ImportError:
        return {}

    tag_rects: dict = {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) \
             else wb.active

        # Build merge lookup: (row, col) → (min_row, min_col, max_row, max_col)
        merge_map: dict = {}
        for mr in ws.merged_cells.ranges:
            for r2 in range(mr.min_row, mr.max_row + 1):
                for c2 in range(mr.min_col, mr.max_col + 1):
                    merge_map[(r2, c2)] = (mr.min_row, mr.min_col,
                                           mr.max_row, mr.max_col)

        max_col = ws.max_column or 50
        max_row = ws.max_row or 100

        # Cumulative x: cum_x[c] = left edge of column c+1 (0-based accumulation)
        # cum_x[0]=0, cum_x[1]=width of col 1, cum_x[2]=width of cols 1+2, ...
        cum_x = [0.0] * (max_col + 2)
        for c in range(1, max_col + 1):
            cum_x[c] = cum_x[c - 1] + _excel_col_width_pt(ws, c)

        # Cumulative y: cum_y[r] = top edge of row r+1
        cum_y = [0.0] * (max_row + 2)
        for r in range(1, max_row + 1):
            cum_y[r] = cum_y[r - 1] + _excel_row_height_pt(ws, r)

        for row in ws.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                text = str(cell.value).strip()
                if '{{' not in text:
                    continue
                r, c = cell.row, cell.column
                min_row, min_col, max_row2, max_col2 = \
                    merge_map.get((r, c), (r, c, r, c))
                # x0 = left edge of min_col,  x1 = right edge of max_col
                x0 = cum_x[min_col - 1]
                x1 = cum_x[max_col2]
                y0 = cum_y[min_row - 1]
                y1 = cum_y[max_row2]
                for m in TAG_TYPED.finditer(text):
                    tag_rects[m.group(0)] = (x0, y0, x1, y1)
                for m in TAG_PLAIN.finditer(text):
                    tag_rects[m.group(0)] = (x0, y0, x1, y1)

        wb.close()
    except Exception as e:
        print(f"⚠️  Could not build Excel tag rects: {e}")

    return tag_rects


def _fit_linear(samples: list) -> tuple:
    """
    Least-squares fit of  y = a*x + b  from [(x,y), ...] pairs.
    Returns (a, b).  With a single sample, returns slope=1 with matching offset.
    Used for the Y-axis (row heights are uniform so linear is accurate).
    """
    n = len(samples)
    if n == 0:
        return 1.0, 0.0
    if n == 1:
        x0, y0 = samples[0]
        return 1.0, y0 - x0
    sx  = sum(x for x, y in samples)
    sy  = sum(y for x, y in samples)
    sxx = sum(x * x for x, y in samples)
    sxy = sum(x * y for x, y in samples)
    denom = n * sxx - sx * sx
    if abs(denom) < 1e-10:
        return 1.0, sy / n - sx / n
    a = (n * sxy - sx * sy) / denom
    b = (sy - a * sx) / n
    return a, b


def _build_piecewise(anchors: list) -> list:
    """
    Build a piecewise-linear X transform from (x_excel, x_pdf) anchor pairs.

    Returns a sorted, deduplicated list of (x_excel, x_pdf) pairs.
    Use _interp_piecewise() to evaluate.

    Why piecewise instead of a single linear scale?
    Excel/LibreOffice column width rendering is NOT perfectly uniform:
    narrow spacer columns (<3 char-units) are rendered proportionally wider
    than the standard MDW formula predicts.  A single linear scale will be
    accurate in some regions and off in others.  A piecewise map built from
    actual PDF border positions avoids this entirely.
    """
    if not anchors:
        return []
    # Deduplicate: keep median pdf_x for each unique excel_x
    from collections import defaultdict
    groups: dict = defaultdict(list)
    for ex, px in anchors:
        groups[round(ex, 1)].append(px)
    result = []
    for ex, pxs in groups.items():
        pxs.sort()
        result.append((ex, pxs[len(pxs) // 2]))  # median
    result.sort()
    return result


def _interp_piecewise(x: float, transform: list) -> float | None:
    """
    Piecewise-linear interpolation / extrapolation.
    transform: sorted list of (x_excel, x_pdf) from _build_piecewise().
    Returns None if transform is empty.
    """
    if not transform:
        return None
    if len(transform) == 1:
        ex0, px0 = transform[0]
        return px0 + (x - ex0)          # unit slope extrapolation
    if x <= transform[0][0]:
        ex0, px0 = transform[0]
        ex1, px1 = transform[1]
        slope = (px1 - px0) / (ex1 - ex0) if ex1 != ex0 else 1.0
        return px0 + (x - ex0) * slope
    if x >= transform[-1][0]:
        ex0, px0 = transform[-2]
        ex1, px1 = transform[-1]
        slope = (px1 - px0) / (ex1 - ex0) if ex1 != ex0 else 1.0
        return px1 + (x - ex1) * slope
    # Binary search for bracket
    lo, hi = 0, len(transform) - 1
    while lo + 1 < hi:
        mid = (lo + hi) // 2
        if transform[mid][0] <= x:
            lo = mid
        else:
            hi = mid
    ex0, px0 = transform[lo]
    ex1, px1 = transform[hi]
    t = (x - ex0) / (ex1 - ex0) if ex1 != ex0 else 0.0
    return px0 + t * (px1 - px0)


def _snap_to_border(span_bbox: tuple, hlines: list) -> tuple | None:
    """
    Given a tag span bbox and all horizontal border lines on the page,
    find the NARROWEST line that:
      - contains the span's x-midpoint (within 2 pt tolerance)
      - is within 3× the span height vertically from the span centre
    Returns (line_x0, line_x1) — the exact cell left/right boundaries.
    Returns None if no suitable line is found.

    Choosing the narrowest line (smallest x-range) avoids mistakenly
    selecting the wide outer page-frame border instead of the cell border.
    """
    sx0, sy0, sx1, sy1 = span_bbox
    smx = (sx0 + sx1) / 2
    smy = (sy0 + sy1) / 2
    row_h = max(sy1 - sy0, 5.0)

    best = None
    best_w = float("inf")
    for y, lx0, lx1 in hlines:
        if abs(y - smy) > row_h * 3:
            continue
        if lx0 - 2 <= smx <= lx1 + 2:
            w = lx1 - lx0
            if 4 < w < best_w:
                best_w = w
                best = (lx0, lx1)
    return best


# ── Font name mapping ────────────────────────────────────────────────────────
# AcroForm fields only support the 14 standard PDF fonts.
# Map common Excel/LibreOffice font names to the closest PDF standard font.
_FONT_MAP = {
    "helv": "Helv", "helvetica": "Helv", "arial": "Helv",
    "arialmt": "Helv", "calibri": "Helv", "tahoma": "Helv",
    "verdana": "Helv", "trebuchet": "Helv", "gill": "Helv",
    "tiro": "TiRo", "times": "TiRo", "timesnewroman": "TiRo",
    "timesnewromanpsmt": "TiRo", "georgia": "TiRo", "palatino": "TiRo",
    "garamond": "TiRo", "bookman": "TiRo",
    "cour": "Cour", "courier": "Cour", "couriernew": "Cour",
    "lucidaconsole": "Cour", "consolasmt": "Cour", "consolas": "Cour",
}

def _pdf_font(span_font: str) -> str:
    """Map a span font name to the nearest standard PDF form font."""
    key = span_font.lower().replace(" ", "").replace("-", "").replace("_", "")
    for suffix in ("bolditalic", "bold", "italic", "oblique", "regular", "mt", "ps"):
        if key.endswith(suffix):
            key = key[: -len(suffix)]
    return _FONT_MAP.get(key, "Helv")


def _infer_align(span_bbox: tuple, cell_rect: fitz.Rect) -> int:
    """
    Infer text alignment (0=left, 1=center, 2=right) from where the tag
    text sits horizontally within the cell rectangle.
    """
    sx0, _, sx1, _ = span_bbox
    cx0, cx1 = cell_rect.x0, cell_rect.x1
    cell_w = cx1 - cx0
    if cell_w < 2:
        return 0
    rel = ((sx0 + sx1) / 2 - cx0) / cell_w   # 0=far left, 1=far right
    if rel < 0.35:
        return 0   # left
    elif rel > 0.65:
        return 2   # right
    else:
        return 1   # center


# ── Excel → PDF ──────────────────────────────────────────────────────────────

def excel_to_pdf(xlsx_path: str, sheet_name: str | None = None) -> bytes:
    """
    Convert an Excel file to PDF bytes.

    Priority:
      1. LibreOffice (headless) — works on ARM, no COM needed
      2. win32com / Microsoft Excel — x64 Windows only

    sheet_name: not used for whole-file export (both engines export all sheets).
    To export a single sheet, hide others beforehand or use --sheet-name post-filter.
    """
    abs_path = os.path.abspath(xlsx_path)
    out_dir  = os.path.dirname(abs_path)

    # ── Try LibreOffice first ────────────────────────────────────────────────
    soffice = _find_libreoffice()
    if soffice:
        return _excel_to_pdf_libreoffice(soffice, abs_path, out_dir, sheet_name)

    # ── Fall back to win32com (Excel COM, x64 only) ──────────────────────────
    return _excel_to_pdf_win32(abs_path, sheet_name)


def _find_libreoffice() -> str | None:
    """Return path to soffice.exe if LibreOffice is installed, else None."""
    import shutil
    # Check PATH first
    found = shutil.which("soffice")
    if found:
        return found
    # Common install locations on Windows
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        os.path.expandvars(r"%LOCALAPPDATA%\Programs\LibreOffice\program\soffice.exe"),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def _excel_to_pdf_libreoffice(soffice: str, abs_path: str, out_dir: str,
                               sheet_name: str | None) -> bytes:
    """Export Excel → PDF via LibreOffice headless."""
    import subprocess

    print(f"🔧  Using LibreOffice: {soffice}")

    # LibreOffice always exports all sheets to the same folder as the input
    result = subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", out_dir, abs_path],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        sys.exit(f"❌  LibreOffice failed:\n{result.stderr}")

    # LibreOffice names the output: <basename>.pdf
    base    = os.path.splitext(os.path.basename(abs_path))[0]
    pdf_out = os.path.join(out_dir, base + ".pdf")

    if not os.path.exists(pdf_out):
        sys.exit(f"❌  LibreOffice ran but output not found: {pdf_out}")

    with open(pdf_out, "rb") as f:
        pdf_bytes = f.read()
    os.unlink(pdf_out)   # clean up intermediate file

    print(f"✅  Excel → PDF via LibreOffice ({len(pdf_bytes):,} bytes)")
    return pdf_bytes


def _excel_to_pdf_win32(abs_path: str, sheet_name: str | None) -> bytes:
    """Export Excel → PDF via win32com (requires Microsoft Excel, x64 only)."""
    try:
        import win32com.client
    except ImportError:
        sys.exit(
            "❌  Cannot convert Excel to PDF — no engine found.\n\n"
            "  Option A (ARM / any Windows):\n"
            "    Install LibreOffice from https://www.libreoffice.org/download/\n\n"
            "  Option B (x64 Windows with Microsoft Excel):\n"
            "    py -m pip install pywin32\n"
        )

    out_dir = os.path.dirname(abs_path)
    tmp_pdf = os.path.join(out_dir, "_tmp_excel_export.pdf")

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(abs_path)
        if sheet_name:
            target = None
            hidden = []
            for ws in wb.Worksheets:
                if ws.Name.lower() == sheet_name.lower():
                    target = ws
                else:
                    if ws.Visible == -1:
                        ws.Visible = 0
                        hidden.append(ws)
            if target is None:
                print(f"⚠️  Sheet '{sheet_name}' not found — exporting all sheets")
                wb.ExportAsFixedFormat(0, tmp_pdf)
            else:
                target.ExportAsFixedFormat(0, tmp_pdf)
            for ws in hidden:
                ws.Visible = -1
        else:
            wb.ExportAsFixedFormat(0, tmp_pdf)
        wb.Close(False)
    finally:
        excel.Quit()

    with open(tmp_pdf, "rb") as f:
        pdf_bytes = f.read()
    os.unlink(tmp_pdf)

    print(f"✅  Excel → PDF via Microsoft Excel ({len(pdf_bytes):,} bytes)")
    return pdf_bytes

def _build_drive_service():
    """
    Returns an authenticated Google Drive v3 service.
    Tries service_account.json first, then Application Default Credentials.
    """
    SA_FILE = os.path.join(os.path.dirname(__file__), "service_account.json")
    SCOPES  = ["https://www.googleapis.com/auth/drive.readonly"]

    try:
        from googleapiclient.discovery import build
    except ImportError:
        sys.exit("❌  Missing google-api-python-client.\n"
                 "    Run:  py -m pip install google-auth google-api-python-client")

    if os.path.exists(SA_FILE):
        from google.oauth2 import service_account
        creds = service_account.Credentials.from_service_account_file(SA_FILE, scopes=SCOPES)
        print(f"🔑  Authenticated via service account: {SA_FILE}")
    else:
        try:
            import google.auth
            creds, _ = google.auth.default(scopes=SCOPES)
            print("🔑  Authenticated via Application Default Credentials")
        except Exception:
            sys.exit(
                "❌  No credentials found.\n"
                "    Option A: Place service_account.json next to this script.\n"
                "    Option B: Run  gcloud auth application-default login"
            )

    from googleapiclient.discovery import build
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def download_sheet_as_pdf(sheet_id: str, sheet_name: str | None = None) -> bytes:
    """
    Export a Google Sheet (or a specific tab) as PDF bytes.

    If sheet_name is given, only that tab is exported.
    The Drive API export always exports all sheets; to isolate one tab we use
    the Sheets API gid parameter via a direct URL request.
    """
    drive = _build_drive_service()

    if sheet_name:
        # Export a specific sheet tab by gid
        # 1. Find the gid of the named sheet via Sheets API
        try:
            from googleapiclient.discovery import build as gapi_build
            from google.oauth2 import service_account

            SA_FILE = os.path.join(os.path.dirname(__file__), "service_account.json")
            SCOPES_SH = ["https://www.googleapis.com/auth/spreadsheets.readonly",
                         "https://www.googleapis.com/auth/drive.readonly"]

            if os.path.exists(SA_FILE):
                creds = service_account.Credentials.from_service_account_file(SA_FILE, scopes=SCOPES_SH)
            else:
                import google.auth
                creds, _ = google.auth.default(scopes=SCOPES_SH)

            sheets = gapi_build("sheets", "v4", credentials=creds, cache_discovery=False)
            meta = sheets.spreadsheets().get(spreadsheetId=sheet_id, fields="sheets.properties").execute()
            gid = None
            for s in meta.get("sheets", []):
                if s["properties"]["title"].lower() == sheet_name.lower():
                    gid = s["properties"]["sheetId"]
                    break
            if gid is None:
                print(f"⚠️  Sheet tab '{sheet_name}' not found — exporting entire spreadsheet")

        except Exception as e:
            print(f"⚠️  Could not look up sheet gid ({e}) — exporting entire spreadsheet")
            gid = None

        if gid is not None:
            # Build export URL with gid filter
            import google.auth.transport.requests
            request = google.auth.transport.requests.Request()
            creds.refresh(request)
            token = creds.token
            url = (
                f"https://docs.google.com/spreadsheets/d/{sheet_id}/export"
                f"?format=pdf&gid={gid}&portrait=true&size=A4"
            )
            import urllib.request
            req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
            with urllib.request.urlopen(req) as resp:
                pdf_bytes = resp.read()
            print(f"✅  Downloaded sheet tab '{sheet_name}' as PDF ({len(pdf_bytes):,} bytes)")
            return pdf_bytes

    # Fallback: export entire spreadsheet as PDF via Drive API
    from googleapiclient.http import MediaIoBaseDownload
    request = drive.files().export_media(fileId=sheet_id, mimeType="application/pdf")
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    pdf_bytes = buf.getvalue()
    print(f"✅  Downloaded spreadsheet as PDF ({len(pdf_bytes):,} bytes)")
    return pdf_bytes


# PDF AcroForm field flag constants (raw integers — compatible with all PyMuPDF versions)
_FLAG_READONLY  = 1       # bit 1
_FLAG_MULTILINE = 4096    # bit 13


def make_text_field(page, rect, name, multiline=False, readonly=False,
                    fontsize=0, fontname="Helv", align=0):
    """
    Add a TextField AcroForm widget.
    fontsize : match Excel cell font size (0 = auto-fit)
    fontname : nearest PDF standard font (Helv / TiRo / Cour)
    align    : 0=left, 1=center, 2=right
    """
    widget = fitz.Widget()
    widget.rect          = rect
    widget.field_type    = fitz.PDF_WIDGET_TYPE_TEXT
    widget.field_name    = name
    widget.field_flags   = (_FLAG_READONLY if readonly else 0) | (_FLAG_MULTILINE if multiline else 0)
    widget.text_fontsize = fontsize
    widget.text_font     = fontname
    widget.text_color    = (0, 0, 0)
    widget.fill_color    = None
    widget.border_width  = 0
    page.add_widget(widget)
    # Set /Q (quadding) for alignment: 0=left 1=center 2=right
    # widget.xref is unreliable on some PyMuPDF builds — find the field by name instead
    if align != 0:
        try:
            for w in page.widgets():
                if w.field_name == name:
                    page.parent.xref_set_key(w.xref, "Q", str(align))
                    break
        except Exception:
            pass


def make_checkbox(page, rect, name):
    """Add a Checkbox AcroForm widget to the page."""
    widget = fitz.Widget()
    widget.rect         = fitz.Rect(rect.x0, rect.y0, rect.x0 + 12, rect.y0 + 12)
    widget.field_type   = fitz.PDF_WIDGET_TYPE_CHECKBOX
    widget.field_name   = name
    widget.field_flags  = 0
    widget.fill_color   = (1, 1, 1)
    widget.border_color = (0.4, 0.4, 0.4)
    widget.border_width = 0.8
    page.add_widget(widget)


def make_sig_box(page, rect, name):
    """Add a visual signature box (plain rectangle, no AcroForm sig field)."""
    page.draw_rect(rect, color=(0.6, 0.6, 0.6), width=0.5, fill=(0.97, 0.97, 0.97))
    widget = fitz.Widget()
    widget.rect         = rect
    widget.field_type   = fitz.PDF_WIDGET_TYPE_TEXT
    widget.field_name   = name
    widget.field_flags  = _FLAG_READONLY
    widget.fill_color   = None
    widget.border_width = 0
    page.add_widget(widget)


def make_photo_box(page, rect, name):
    """
    Draw a visible image-placeholder box.
    No AcroForm widget — the PWA replaces this area with a real image at
    fill time by reading the rect from the field-map JSON.
    A dashed grey border + mountain-and-sun icon makes the area visible.
    """
    # Outer border
    page.draw_rect(rect, color=(0.6, 0.6, 0.6), width=0.6,
                   fill=(0.95, 0.95, 0.95), dashes="[3 2]"
    )
    # Simple icon: circle (sun) + triangle (mountain) centred in the box
    cx = (rect.x0 + rect.x1) / 2
    cy = (rect.y0 + rect.y1) / 2
    r  = min(rect.width, rect.height) * 0.18   # scale to box size
    # sun circle
    page.draw_circle(fitz.Point(cx - r * 0.6, cy - r * 0.5), r * 0.4,
                     color=(0.7, 0.7, 0.7), width=0.5)
    # mountain triangle
    pts = [fitz.Point(cx - r, cy + r),
           fitz.Point(cx,     cy - r * 0.3),
           fitz.Point(cx + r, cy + r)]
    page.draw_polyline(pts + [pts[0]], color=(0.7, 0.7, 0.7), width=0.5)
    # Label — insert_text() has no align param; approximate centre manually
    font_sz  = max(6, min(8, rect.height * 0.12))
    try:
        tw = fitz.get_text_length(name, fontname="helv", fontsize=font_sz)
    except Exception:
        tw = font_sz * len(name) * 0.5   # rough fallback
    page.insert_text(
        fitz.Point(cx - tw / 2, rect.y1 - 3),
        name,
        fontsize=font_sz,
        color=(0.6, 0.6, 0.6),
    )


def find_cell_rect(span_bbox: tuple, drawings: list,
                   multiline: bool = False, photo: bool = False) -> fitz.Rect:
    """
    Fallback rect finder used when no Excel geometry data is available.
    Picks the tightest drawn filled rect whose midpoint contains the span midpoint
    (works for LibreOffice exports that render cell fills as vector rects).
    Falls back to the padded span bbox if no filled rect is found.
    """
    sx0, sy0, sx1, sy1 = span_bbox
    mx = (sx0 + sx1) / 2
    my = (sy0 + sy1) / 2

    best_rect = None
    best_area = float("inf")
    for d in drawings:
        r = d.get("rect")
        if r is None:
            continue
        if r.x0 - 1 <= mx <= r.x1 + 1 and r.y0 - 1 <= my <= r.y1 + 1:
            area = (r.x1 - r.x0) * (r.y1 - r.y0)
            if area < best_area and area > 4:
                best_area = area
                best_rect = r

    if best_rect is not None:
        return fitz.Rect(best_rect.x0 + 1, best_rect.y0 + 1,
                         best_rect.x1 - 1, best_rect.y1 - 1)

    # Fallback: padded span bbox
    pad = 2
    return fitz.Rect(sx0 - pad, sy0 - pad, sx1 + pad, sy1 + pad)


def convert(input_path: str, output_path: str, excel_tags: set = None,
            excel_tag_rects: dict = None):
    doc = fitz.open(input_path)
    try:
        _do_convert(doc, output_path,
                    excel_tags=excel_tags or set(),
                    excel_tag_rects=excel_tag_rects or {})
    except Exception:
        doc.close()
        raise


def scan_excel_tags(xlsx_path: str, sheet_name: str = None) -> set:
    """
    Read every cell in the Excel workbook and return the set of all complete
    tag strings found (e.g. '{{date:confirmedDate}}', '{{ml:title:3}}').

    This is used as a fallback when a PDF span contains only a truncated
    fragment of a tag (because the Excel cell was too narrow to display the
    full text and LibreOffice only embedded the visible portion in the PDF).
    """
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl not installed — narrow-cell tag recovery disabled.")
        print("   Run: py -m pip install openpyxl")
        return set()

    tags: set = set()
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheets = [wb[sheet_name]] if (sheet_name and sheet_name in wb.sheetnames) \
                 else wb.worksheets
        for ws in sheets:
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if not val:
                        continue
                    text = str(val).strip()
                    if '{{' not in text:
                        continue
                    for m in TAG_TYPED.finditer(text):
                        tags.add(m.group(0))
                    for m in TAG_PLAIN.finditer(text):
                        tags.add(m.group(0))
        wb.close()
    except Exception as e:
        print(f"⚠️  Could not scan Excel tags: {e}")

    return tags


def _do_convert(doc, output_path: str, excel_tags: set = None,
                excel_tag_rects: dict = None):
    """
    Two-pass page converter.

    Pass 1 — scan every page for tag spans.  For each tag that also exists in
    excel_tag_rects, record its PDF span position as a calibration sample so
    we can derive the linear mapping:
        x_pdf = xscale * x_excel + xoff
        y_pdf = yscale * y_excel + yoff

    Pass 2 — using the calibrated transform, compute each field's full
    bounding rect from the Excel merge geometry and place the AcroForm widget.
    Falls back to find_cell_rect (PDF-drawing heuristic) when Excel data is
    unavailable (e.g. Mode C: PDF-only input).
    """
    field_map = {}   # { fieldName: { type, page, rect } }
    excel_tags      = excel_tags or set()
    excel_tag_rects = excel_tag_rects or {}

    for page_num, page in enumerate(doc):
        drawings = page.get_drawings()

        # ── Pass 1: collect all tag spans on this page ────────────────────────
        # Each entry: (tag_text, parsed_tuple, span_dict)
        all_spans: list = []

        for block in page.get_text("dict")["blocks"]:
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = span["text"].strip()
                    parsed = parse_tag(text)

                    # Narrow-cell recovery: LibreOffice sometimes clips tag text
                    # to the visible cell width.  Recover the full tag string.
                    #
                    # Guard to prevent false matches on static label words:
                    #   "Page" ⊂ "{{totalPage:c}}"  — must NOT fire (label word)
                    #   "alPage" ⊂ "{{totalPage:c}}" — must fire (real clip)
                    #
                    # A span qualifies for recovery if it looks like a tag fragment:
                    #   (a) contains a tag-syntax char  { } :   — definitely a clip
                    #   (b) OR is a pure identifier (alphanumeric+_) starting with a
                    #       lowercase letter — camelCase field-name fragment.
                    #       Static labels always start with uppercase or contain
                    #       spaces / punctuation, so they are excluded by this rule.
                    _is_tag_fragment = (
                        any(c in text for c in '{}:')
                        or (len(text) >= 3
                            and text[0].islower()
                            and text.replace('_', '').isalnum())
                    )
                    if parsed is None and len(text) >= 3 and excel_tags and _is_tag_fragment:
                        for full_tag in excel_tags:
                            if text in full_tag:
                                parsed = parse_tag(full_tag)
                                if parsed:
                                    text = full_tag
                                    break

                    if not parsed:
                        continue

                    all_spans.append((text, parsed, span))

        # ── Build horizontal-border-line index for this page ─────────────────
        # Used as calibration anchors: each tag's cell has a top/bottom border
        # line in the PDF that gives the EXACT cell x0 and x1.
        # We choose the NARROWEST line containing the span midpoint to avoid
        # picking the wide outer page-frame border.
        hlines_page: list = []   # [(y, x0, x1), ...]
        for d in drawings:
            r = d.get("rect")
            if r and abs(r.y0 - r.y1) < 1.5:
                hlines_page.append((r.y0, r.x0, r.x1))

        # ── Calibrate Excel → PDF X transform (piecewise) ────────────────────
        # For each tag whose Excel cell rect is known:
        #   1. Snap to the actual PDF cell border via _snap_to_border()
        #      → gives exact (border_x0, border_x1) in PDF space
        #   2. Add (exc_x0, border_x0) and (exc_x1, border_x1) as anchors
        # Fallback (no border found): alignment-aware span-position anchor.
        #
        # Why piecewise?  Narrow spacer columns (<3 char-units) render
        # proportionally wider than the MDW formula predicts, making the
        # Excel→PDF mapping non-linear.  A piecewise map from actual border
        # positions is exact at every sampled column boundary.
        cal_x_anchors: list = []   # (x_excel, x_pdf) pairs for piecewise X
        cal_y:         list = []   # (y_excel_mid, y_pdf_mid) for linear Y

        for tag_text, parsed, span in all_spans:
            if tag_text not in excel_tag_rects:
                continue
            _, _, _, _, explicit_align = parsed
            ex0, ey0, ex1, ey1 = excel_tag_rects[tag_text]
            sb = span["bbox"]

            # ── X: snap to cell border ───────────────────────────────────
            snapped = _snap_to_border(sb, hlines_page)
            if snapped is not None:
                border_x0, border_x1 = snapped
                cal_x_anchors.append((ex0, border_x0))
                cal_x_anchors.append((ex1, border_x1))
            else:
                # No border line found — fall back to alignment-based span pos
                if explicit_align == 1:        # center
                    cal_x_anchors.append(((ex0 + ex1) / 2, (sb[0] + sb[2]) / 2))
                elif explicit_align == 2:      # right
                    cal_x_anchors.append((ex1, sb[2]))
                else:                          # left (explicit or auto)
                    cal_x_anchors.append((ex0, sb[0]))

            # ── Y: span vertical centre (text is always vertically centred) ─
            cal_y.append(((ey0 + ey1) / 2, (sb[1] + sb[3]) / 2))

        x_transform  = _build_piecewise(cal_x_anchors)
        have_x_transform = len(x_transform) >= 2
        have_y_transform = len(cal_y) >= 2
        if have_y_transform:
            yscale, yoff = _fit_linear(cal_y)
        else:
            yscale = yoff = None
        have_calibration = have_x_transform and have_y_transform
        if not have_calibration and excel_tag_rects:
            print(f"⚠️  Not enough calibration samples "
                  f"(x_anchors={len(cal_x_anchors)}, y={len(cal_y)}) — "
                  "falling back to PDF-drawing rect detection")

        # ── Pass 2: place AcroForm widgets ────────────────────────────────────
        # Photo boxes must be drawn AFTER apply_redactions() to survive it.
        pending_photos: list = []

        for tag_text, parsed, span in all_spans:
            tag_type, tag_name, lines, spacing, explicit_align = parsed
            span_fontsize = span.get("size", 0)
            span_font     = _pdf_font(span.get("font", "Helv"))
            span_bbox     = span["bbox"]

            # ── Compute field rect ───────────────────────────────────────────
            if have_calibration and tag_text in excel_tag_rects:
                # Primary path: piecewise X transform from actual border lines
                # + linear Y transform from span centres.
                # This handles non-linear column width rendering accurately.
                ex0, ey0, ex1, ey1 = excel_tag_rects[tag_text]
                rx0 = _interp_piecewise(ex0, x_transform)
                rx1 = _interp_piecewise(ex1, x_transform)
                ry0 = yscale * ey0 + yoff
                ry1 = yscale * ey1 + yoff
                rect = fitz.Rect(rx0 + 1, ry0 + 1, rx1 - 1, ry1 - 1)
            else:
                # Fallback: infer from PDF drawing paths (filled rects).
                rect = find_cell_rect(span_bbox, drawings,
                                      multiline=(tag_type == "ml"),
                                      photo=(tag_type == "photo"))

            # {{ml:name:N}} — override height to exactly N lines
            if tag_type == "ml" and lines:
                fs     = span_fontsize if span_fontsize > 0 else 10
                line_h = fs * (spacing if spacing is not None else _VIEWER_LINE_HEIGHT)
                sx0, sy0 = span_bbox[0], span_bbox[1]
                rect = fitz.Rect(rect.x0, sy0 - 2,
                                 rect.x1, sy0 - 2 + lines * line_h + _VIEWER_LINE_PADDING)

            # Redact the tag text (transparent — preserves borders / background)
            sx0r, sy0r, sx1r, sy1r = span_bbox
            page.add_redact_annot(
                fitz.Rect(sx0r, sy0r, sx1r, sy1r),
                fill=None,
                cross_out=False
            )

            align = explicit_align if explicit_align is not None \
                    else _infer_align(span_bbox, rect)

            field_map[tag_name] = {
                "type":        tag_type,
                "page":        page_num,
                "rect":        [rect.x0, rect.y0, rect.x1, rect.y1],
                "fontsize":    round(span_fontsize, 1),
                "font":        span_font,
                "align":       align,
                "lines":       lines,
                "lineSpacing": spacing
            }

            if tag_type == "txt":
                make_text_field(page, rect, tag_name,
                                fontsize=span_fontsize, fontname=span_font, align=align)
            elif tag_type == "ml":
                make_text_field(page, rect, tag_name, multiline=True,
                                fontsize=span_fontsize, fontname=span_font, align=align)
            elif tag_type == "chk":
                make_checkbox(page, rect, tag_name)
            elif tag_type == "sig":
                make_sig_box(page, rect, tag_name)
            elif tag_type == "photo":
                pending_photos.append((rect, tag_name))
            elif tag_type == "date":
                make_text_field(page, rect, tag_name,
                                fontsize=span_fontsize, fontname=span_font, align=align)

        # Apply all redactions (removes tag text spans only, preserves borders)
        page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)

        # Draw photo placeholder boxes NOW — after redactions so they survive
        for p_rect, p_name in pending_photos:
            make_photo_box(page, p_rect, p_name)

    doc.save(output_path, garbage=4, deflate=True)
    doc.close()
    doc = None

    # Save field map alongside the output
    map_path = output_path.replace(".pdf", "_fieldmap.json")
    with open(map_path, "w") as f:
        json.dump(field_map, f, indent=2)

    print(f"✅ Saved: {output_path}")
    print(f"✅ Field map: {map_path}")
    print(f"   Fields found: {len(field_map)}")
    for name, info in field_map.items():
        print(f"   [{info['type']:4s}] {name}  (page {info['page'] + 1})")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert Excel / Google Sheet / PDF with {{tags}} into a fillable AcroForm PDF."
    )
    parser.add_argument("input", nargs="?", default=None,
                        help="Input file: .xlsx/.xls (Excel) or .pdf (Mode B/C)")
    parser.add_argument("output", nargs="?", default=None,
                        help="Output PDF path, e.g. itr_cover_form.pdf")
    parser.add_argument("--sheet-id", metavar="ID",
                        help="Google Sheets ID to download directly (Mode B)")
    parser.add_argument("--sheet-name", metavar="NAME", default=None,
                        help="Sheet tab name to export (e.g. 'cover'). Omit to export all tabs.")
    args = parser.parse_args()

    # When --sheet-id is used, the single positional arg is the output path
    if args.sheet_id and args.input and args.output is None:
        args.output = args.input
        args.input  = None

    if not args.output:
        parser.print_help()
        sys.exit(1)

    def _run_from_bytes(pdf_bytes: bytes, out_path: str,
                         excel_tags: set = None, excel_tag_rects: dict = None):
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.write(pdf_bytes)
        tmp.close()
        try:
            convert(tmp.name, out_path,
                    excel_tags=excel_tags, excel_tag_rects=excel_tag_rects)
        except Exception as e:
            print(f"\u274c  Conversion error: {e}")
            raise
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass   # already gone or still locked — ignore

    if args.input and args.input.lower().endswith((".xlsx", ".xls")):
        # Mode A: Excel → PDF → convert
        # Scan full tag strings AND exact merge-range rects BEFORE conversion.
        # Tags set: used to recover narrow-cell truncated tag strings in the PDF.
        # Tag rects: used to place AcroForm fields at the correct merged-cell size.
        excel_tags      = scan_excel_tags(args.input, args.sheet_name)
        excel_tag_rects = build_excel_tag_rects(args.input, args.sheet_name)
        if excel_tags:
            print(f"\U0001f4cb  Excel tags scanned: {len(excel_tags)} tag(s) found")
        pdf_bytes = excel_to_pdf(args.input, args.sheet_name)
        _run_from_bytes(pdf_bytes, args.output,
                        excel_tags=excel_tags, excel_tag_rects=excel_tag_rects)

    elif args.sheet_id:
        # Mode B: Google Sheet → PDF → convert
        pdf_bytes = download_sheet_as_pdf(args.sheet_id, args.sheet_name)
        _run_from_bytes(pdf_bytes, args.output)

    elif args.input and args.input.lower().endswith(".pdf"):
        # Mode C: local PDF
        convert(args.input, args.output)

    else:
        parser.print_help()
        sys.exit(1)
